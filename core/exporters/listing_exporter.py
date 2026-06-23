from __future__ import annotations

import csv
import html
import base64
import unicodedata
from datetime import datetime
from pathlib import Path

from core.db import Project
from core.exporters.case_context import build_case_context, case_export_metadata_rows, context_cards_html
from core.formatters import format_short_date

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_EXCEL_MAX_CELL_LEN = 32767
_EXCEL_MAX_SHEET_TITLE_LEN = 31


def _sanitize_excel_sheet_title(title: str) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", str(title or ""))
    text = "".join(ch for ch in text if ch not in "[]:*?/\\'")
    text = text.strip() or "Export"
    return text[:_EXCEL_MAX_SHEET_TITLE_LEN] or "Export"


def _sanitize_excel_cell(value: object) -> str:
    if value is None:
        return ""
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    cleaned: list[str] = []
    for ch in text:
        code = ord(ch)
        if unicodedata.category(ch) in {"Cc", "Cs"}:
            continue
        if code in {0xFEFF} or 0xE0000 <= code <= 0xE007F:
            continue
        cleaned.append(ch)
    text = "".join(cleaned)
    if len(text) > _EXCEL_MAX_CELL_LEN:
        return text[: _EXCEL_MAX_CELL_LEN - 3] + "..."
    return text


def export_listing_csv(
    file_path: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    project: Project | None = None,
    project_name: str = "",
    dataset_meta: dict | None = None,
) -> None:
    path = Path(file_path)
    case_context = build_case_context(project, project_name=project_name, dataset_meta=dataset_meta)
    metadata_rows = case_export_metadata_rows(case_context)

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for label, value in metadata_rows:
            writer.writerow([f"# {label}: {value}"])
        if metadata_rows:
            writer.writerow([])
        writer.writerow(headers)
        writer.writerows(rows)

def export_listing_excel(
    file_path: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    sheet_title: str = "Listing",
    project: Project | None = None,
    project_name: str = "",
    dataset_meta: dict | None = None,
) -> None:
    path = Path(file_path)
    case_context = build_case_context(project, project_name=project_name, dataset_meta=dataset_meta)
    metadata_rows = case_export_metadata_rows(case_context)

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_excel_sheet_title(sheet_title)

    meta_font = Font(bold=True, color="374151")
    for label, value in metadata_rows:
        ws.append([_sanitize_excel_cell(label), _sanitize_excel_cell(value)])
        label_cell = ws.cell(row=ws.max_row, column=1)
        label_cell.font = meta_font
    if metadata_rows:
        ws.append([])

    safe_headers = [_sanitize_excel_cell(header) for header in headers]
    header_row = ws.max_row + 1
    ws.append(safe_headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(safe_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([_sanitize_excel_cell(cell) for cell in row])

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Autosize columns
    for col_idx, header in enumerate(safe_headers, start=1):
        max_len = len(str(header))

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb.save(path)

def _load_listing_html_template() -> str:
    project_root = Path(__file__).resolve().parents[2]
    template_path = project_root / "templates" / "listing_export.html"
    return template_path.read_text(encoding="utf-8")

def export_listing_html(
    file_path: str,
    headers: list[str],
    rows: list[list[str]],
    dataset: str,
    view_mode: str,
    files_count: int,
    meta: dict | None = None,
    project: Project | None = None,
    project_name: str = "",
) -> None:
    path = Path(file_path)
    text = _report_text()

    meta = meta or {}

    klasa = str(meta.get("OrigRegNo") or "-")
    urbroj = str(meta.get("RegNo") or "-")
    target = str(meta.get("target") or "-")
    targettype = str(meta.get("targettype") or "")

    target_display = target
    case_context = build_case_context(project, project_name=project_name, dataset_meta=meta)

    bt = format_short_date(meta.get("bt"), missing="-")
    et = format_short_date(meta.get("et"), missing="-")

    period = "-"
    if bt != "-" or et != "-":
        period = f"{bt} – {et}"

    dataset_name = Path(dataset).name if dataset else "(no dataset)"
    project_root = Path(__file__).resolve().parents[2]
    logo_path = project_root / "assets" / "ViaNyquist.png"

    logo_data_uri = ""
    if logo_path.exists():
        logo_b64 = base64.b64encode(logo_path.read_bytes()).decode("ascii")
        logo_data_uri = f"data:image/png;base64,{logo_b64}"

    template = _load_listing_html_template()

    table_headers = "".join(
        f"<th>{html.escape(str(header))}</th>"
        for header in headers
    )

    table_rows_parts = []
    for row in rows:
        cells = "".join(
            f"<td>{html.escape(str(cell))}</td>"
            for cell in row
        )
        table_rows_parts.append(f"<tr>{cells}</tr>")

    table_rows = "\n".join(table_rows_parts)

    rendered = (
        template
        .replace("{{LANG}}", "en")
        .replace("{{TITLE}}", html.escape(text["title"]))
        .replace("{{REPORT_TITLE}}", html.escape(text["title"]))
        .replace("{{LOGO}}", html.escape(logo_data_uri))
        .replace("{{DATASET}}", html.escape(dataset_name))
        .replace("{{DATASET_LABEL}}", html.escape(text["dataset"]))
        .replace("{{EXPORTED_AT}}", datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
        .replace("{{EXPORTED_LABEL}}", html.escape(text["exported"]))
        .replace("{{VIEW_MODE}}", html.escape(view_mode or "Unknown"))
        .replace("{{VIEW_LABEL}}", html.escape(text["view"]))
        .replace("{{CASE_CONTEXT_CARDS}}", context_cards_html(case_context, card_class="info"))
        .replace("{{KLASA}}", html.escape(klasa))
        .replace("{{URBROJ}}", html.escape(urbroj))
        .replace("{{TARGET_LABEL}}", html.escape(text["target"]))
        .replace("{{TARGET}}", html.escape(target_display))
        .replace("{{ORDER_VALIDITY_LABEL}}", html.escape(text["order_validity"]))
        .replace("{{PERIOD}}", html.escape(period))
        .replace("{{ROWS_COUNT}}", str(len(rows)))
        .replace("{{ROWS_LABEL}}", html.escape(text["rows"]))
        .replace("{{COLUMNS_COUNT}}", str(len(headers)))
        .replace("{{COLUMNS_LABEL}}", html.escape(text["columns"]))
        .replace("{{FILES_COUNT}}", str(files_count))
        .replace("{{JSON_FILES_LABEL}}", html.escape(text["json_files"]))
        .replace("{{TABLE_TITLE}}", html.escape(text["table_title"]))
        .replace("{{TABLE_HEADERS}}", table_headers)
        .replace("{{TABLE_ROWS}}", table_rows)
    )

    path.write_text(rendered, encoding="utf-8")


def _report_text() -> dict[str, str]:
    return {
        "title": "ViaNyquist Listing Report",
        "dataset": "Dataset",
        "exported": "Exported",
        "view": "View",
        "rows": "Rows",
        "columns": "Columns",
        "json_files": "JSON files",
        "target": "Target",
        "order_validity": "Order validity",
        "table_title": "Listing Data",
    }
