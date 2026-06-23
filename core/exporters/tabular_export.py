from __future__ import annotations

import csv
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_EXCEL_MAX_CELL_LEN = 32767
_EXCEL_MAX_SHEET_TITLE_LEN = 31
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_META_LABEL_FONT = Font(bold=True, color="374151")
_META_HEADER_FONT = Font(bold=True, color="6B7280")


def _sanitize_excel_sheet_title(title: str) -> str:
    text = ILLEGAL_CHARACTERS_RE.sub("", str(title or ""))
    text = "".join(ch for ch in text if ch not in "[]:*?/\\'")
    text = text.strip() or "Export"
    return text[:_EXCEL_MAX_SHEET_TITLE_LEN] or "Export"


def _sanitize_excel_cell(value: object) -> str:
    if value is None:
        return ""
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    cleaned: list[str] = []
    for ch in text:
        code = ord(ch)
        if unicodedata.category(ch) in {"Cc", "Cs"}:
            continue
        if code in {0xFEFF} or 0xE0000 <= code <= 0xE007F:
            continue
        cleaned.append(ch)
    text = "".join(cleaned)
    if len(text) > _EXCEL_MAX_CELL_LEN:
        return text[: _EXCEL_MAX_CELL_LEN - 3] + "..."
    return text


def _autosize_columns(ws, *, min_col: int = 1, max_col: int | None = None, header_row: int = 1) -> None:
    if max_col is None:
        max_col = ws.max_column
    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        for row_idx in range(header_row, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 52)


def _style_header_row(ws, row_num: int, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_tabular_csv(
    file_path: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    metadata_rows: list[tuple[str, str]] | None = None,
) -> None:
    """Write column headers on row 1 so CSV opens cleanly in Excel; case metadata goes after the data."""
    path = Path(file_path)
    safe_headers = [str(header) for header in headers]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(safe_headers)
        writer.writerows(rows)
        metadata_rows = metadata_rows or []
        if metadata_rows:
            writer.writerow([])
            writer.writerow(["Field", "Value"])
            writer.writerows(metadata_rows)


def export_tabular_excel(
    file_path: str,
    headers: list[str],
    rows: list[list[str]],
    *,
    sheet_title: str = "Export",
    metadata_rows: list[tuple[str, str]] | None = None,
) -> None:
    """One worksheet: data table first (frozen header row), case metadata appended below."""
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_excel_sheet_title(sheet_title)

    safe_headers = [_sanitize_excel_cell(header) for header in headers]
    header_row = 1
    ws.append(safe_headers)
    _style_header_row(ws, header_row, len(safe_headers))

    for row in rows:
        ws.append([_sanitize_excel_cell(cell) for cell in row])

    if safe_headers:
        last_col = get_column_letter(len(safe_headers))
        data_last_row = ws.max_row
        ws.auto_filter.ref = f"A{header_row}:{last_col}{data_last_row}"

    ws.freeze_panes = "A2"
    _autosize_columns(ws, header_row=header_row)

    metadata_rows = metadata_rows or []
    if metadata_rows:
        meta_start = ws.max_row + 2
        ws.cell(row=meta_start, column=1, value="Case information")
        ws.cell(row=meta_start, column=1).font = _META_HEADER_FONT
        meta_header_row = meta_start + 1
        ws.cell(row=meta_header_row, column=1, value="Field")
        ws.cell(row=meta_header_row, column=2, value="Value")
        for col in (1, 2):
            cell = ws.cell(row=meta_header_row, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
            cell.font = Font(bold=True, color="111827")
        for offset, (label, value) in enumerate(metadata_rows, start=1):
            row_num = meta_header_row + offset
            ws.cell(row=row_num, column=1, value=_sanitize_excel_cell(label)).font = _META_LABEL_FONT
            ws.cell(row=row_num, column=2, value=_sanitize_excel_cell(value))
        _autosize_columns(ws, min_col=1, max_col=2, header_row=meta_header_row)

    wb.save(Path(file_path))
